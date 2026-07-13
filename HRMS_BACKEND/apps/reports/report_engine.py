"""
Report generation engine for all HRMS modules.
Supports CSV, Excel, and PDF output formats.
"""
import csv
import io
from datetime import date

from django.db.models import Count, Q, Sum, Avg
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
)

from apps.employees.models import Employee, Department, Grade, Position
from apps.leave.models import LeaveRequest, LeaveType
from apps.recruitment.models import JobPosting, JobApplication
from apps.evaluation.models import PerformanceCycle, PerformanceReview, KPI
from apps.transfers.models import Transfer
from apps.manpower.models import ManpowerPlan, EstablishmentPost
from apps.trainees.models import TrainingProgram, Trainee, TraineeAssessment


# ── Helpers ──────────────────────────────────────────────────────────────────

def _apply_excel_styles(ws, headers, col_count):
    """Apply header styling and auto-width to an Excel worksheet."""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    # Auto-width
    for col_idx in range(1, col_count + 1):
        max_len = len(str(ws.cell(row=1, column=col_idx).value or ""))
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2, max_row=ws.max_row):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)


def _build_pdf_table(data, headers, title):
    """Build a PDF document with a title and table."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), title=title)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle", parent=styles["Title"], fontSize=16, spaceAfter=20
    )
    elements = [Paragraph(title, title_style), Spacer(1, 12)]

    table_data = [headers]
    for row in data:
        table_data.append([str(cell) if cell is not None else "" for cell in row])

    col_count = len(headers)
    page_width = landscape(A4)[0] - 72  # minus margins
    col_width = page_width / col_count

    table = Table(table_data, colWidths=[col_width] * col_count, repeatRows=1)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5496")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
    ]
    table.setStyle(TableStyle(style_cmds))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    return buf


# ── Data extractors ──────────────────────────────────────────────────────────

def _headcount_data(params):
    qs = Employee.objects.all()
    if params.get("department_id"):
        qs = qs.filter(department_id=params["department_id"])
    if params.get("gender"):
        qs = qs.filter(gender=params["gender"])
    if params.get("contract_type"):
        qs = qs.filter(contract_type=params["contract_type"])
    if params.get("employment_status"):
        qs = qs.filter(employment_status=params["employment_status"])
    else:
        qs = qs.filter(employment_status="active")

    headers = ["Employee ID", "Full Name", "Department", "Position", "Grade",
               "Gender", "Contract Type", "Status", "Join Date"]
    data = []
    for emp in qs.select_related("user", "department", "position", "grade"):
        data.append([
            emp.employee_id, emp.user.full_name, emp.department.name,
            emp.position.title, emp.grade.level if emp.grade else "",
            emp.get_gender_display(), emp.get_contract_type_display(),
            emp.get_employment_status_display(), emp.join_date,
        ])
    return headers, data


def _leave_data(params):
    qs = LeaveRequest.objects.all()
    if params.get("year"):
        qs = qs.filter(start_date__year=params["year"])
    if params.get("leave_type_id"):
        qs = qs.filter(leave_type_id=params["leave_type_id"])
    if params.get("status"):
        qs = qs.filter(status=params["status"])
    if params.get("department_id"):
        qs = qs.filter(employee__department_id=params["department_id"])

    headers = ["Employee", "Leave Type", "Start Date", "End Date",
               "Days", "Status", "Reason"]
    data = []
    for lr in qs.select_related("employee__user", "leave_type"):
        days = (lr.end_date - lr.start_date).days + 1 if lr.end_date and lr.start_date else 0
        data.append([
            lr.employee.user.full_name, lr.leave_type.name,
            lr.start_date, lr.end_date, days,
            lr.get_status_display(), lr.reason,
        ])
    return headers, data


def _recruitment_data(params):
    qs = JobPosting.objects.all()
    if params.get("status"):
        qs = qs.filter(status=params["status"])
    if params.get("job_type"):
        qs = qs.filter(job_type=params["job_type"])

    headers = ["Job Title", "Department", "Job Type", "Status",
               "Openings", "Applications", "Posted Date", "Closing Date"]
    data = []
    for job in qs.select_related("department"):
        app_count = JobApplication.objects.filter(job=job).count()
        data.append([
            job.title, job.department.name, job.job_type,
            job.get_status_display(), job.openings, app_count,
            job.posted_date, job.closing_date,
        ])
    return headers, data


def _performance_data(params):
    qs = PerformanceReview.objects.all()
    if params.get("cycle_id"):
        qs = qs.filter(cycle_id=params["cycle_id"])
    if params.get("department_id"):
        qs = qs.filter(employee__department_id=params["department_id"])
    if params.get("status"):
        qs = qs.filter(status=params["status"])

    headers = ["Employee", "Department", "Cycle", "Overall Score",
               "Grade", "Status", "Updated"]
    data = []
    for rv in qs.select_related("employee__user", "employee__department", "cycle"):
        data.append([
            rv.employee.user.full_name, rv.employee.department.name,
            rv.cycle.name, rv.overall_score, rv.grade,
            rv.get_status_display(), rv.updated_at.date(),
        ])
    return headers, data


def _transfer_data(params):
    qs = Transfer.objects.all()
    if params.get("transfer_type"):
        qs = qs.filter(transfer_type=params["transfer_type"])
    if params.get("status"):
        qs = qs.filter(status=params["status"])
    if params.get("department_id"):
        qs = qs.filter(
            Q(from_department_id=params["department_id"]) |
            Q(to_department_id=params["department_id"])
        )

    headers = ["Employee", "Type", "From Dept", "To Dept",
               "Effective Date", "Status", "Reason"]
    data = []
    for t in qs.select_related("employee__user", "from_department", "to_department"):
        data.append([
            t.employee.user.full_name, t.get_transfer_type_display(),
            t.from_department.name, t.to_department.name,
            t.effective_date, t.get_status_display(), t.reason,
        ])
    return headers, data


def _manpower_data(params):
    qs = ManpowerPlan.objects.all()
    if params.get("department_id"):
        qs = qs.filter(department_id=params["department_id"])
    if params.get("financial_year"):
        qs = qs.filter(financial_year=params["financial_year"])
    if params.get("status"):
        qs = qs.filter(status=params["status"])

    headers = ["Title", "Department", "Financial Year", "Current Staff",
               "Required Staff", "Gap", "Budget", "Status"]
    data = []
    for mp in qs.select_related("department"):
        data.append([
            mp.title, mp.department.name, mp.financial_year,
            mp.current_staff, mp.required_staff, mp.gap,
            mp.budget, mp.get_status_display(),
        ])
    return headers, data


def _trainee_data(params):
    qs = Trainee.objects.all()
    if params.get("program_id"):
        qs = qs.filter(program_id=params["program_id"])
    if params.get("department_id"):
        qs = qs.filter(department_id=params["department_id"])
    if params.get("status"):
        qs = qs.filter(status=params["status"])
    if params.get("trainee_type"):
        qs = qs.filter(trainee_type=params["trainee_type"])

    headers = ["Name", "Program", "Department", "Type", "Institution",
               "Start Date", "End Date", "Status", "Stipend"]
    data = []
    for t in qs.select_related("user", "program", "department"):
        data.append([
            t.user.full_name, t.program.title, t.department.name,
            t.get_trainee_type_display(), t.institution,
            t.start_date, t.end_date, t.get_status_display(), t.stipend,
        ])
    return headers, data


# ── Format generators ────────────────────────────────────────────────────────

def generate_csv(headers, data):
    """Return CSV bytes."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in data:
        writer.writerow(row)
    return output.getvalue().encode("utf-8")


def generate_excel(headers, data, title):
    """Return Excel bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name limit
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    for row_idx, row in enumerate(data, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    _apply_excel_styles(ws, headers, len(headers))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generate_pdf(headers, data, title):
    """Return PDF bytes."""
    buf = _build_pdf_table(data, headers, title)
    return buf.read()


# ── Main entry point ─────────────────────────────────────────────────────────

REPORT_GENERATORS = {
    "headcount":   _headcount_data,
    "leave":       _leave_data,
    "recruitment": _recruitment_data,
    "performance": _performance_data,
    "transfer":    _transfer_data,
    "manpower":    _manpower_data,
    "trainee":     _trainee_data,
}

REPORT_LABELS = {
    "headcount":   "Headcount Report",
    "leave":       "Leave Summary",
    "recruitment": "Recruitment Report",
    "performance": "Performance Report",
    "transfer":    "Transfer Report",
    "manpower":    "Manpower Report",
    "trainee":     "Trainee Report",
}

REPORT_PARAMETERS = {
    "headcount": [
        {"key": "department_id",    "label": "Department",      "type": "select", "source": "departments"},
        {"key": "gender",           "label": "Gender",          "type": "select", "options": [("M","Male"),("F","Female"),("O","Other")]},
        {"key": "contract_type",    "label": "Contract Type",   "type": "select", "options": [("permanent","Permanent"),("contract","Contract"),("casual","Casual"),("graduate","Graduate Trainee"),("intern","Intern")]},
        {"key": "employment_status","label": "Status",          "type": "select", "options": [("active","Active"),("on_leave","On Leave"),("probation","Probation"),("terminated","Terminated")]},
    ],
    "leave": [
        {"key": "year",             "label": "Year",            "type": "number"},
        {"key": "leave_type_id",    "label": "Leave Type",      "type": "select", "source": "leave_types"},
        {"key": "status",           "label": "Status",          "type": "select", "options": [("pending","Pending"),("approved","Approved"),("rejected","Rejected"),("cancelled","Cancelled")]},
        {"key": "department_id",    "label": "Department",      "type": "select", "source": "departments"},
    ],
    "recruitment": [
        {"key": "status",           "label": "Job Status",      "type": "select", "options": [("open","Open"),("closed","Closed"),("draft","Draft")]},
        {"key": "job_type",         "label": "Job Type",        "type": "select", "options": [("full_time","Full Time"),("part_time","Part Time"),("contract","Contract")]},
    ],
    "performance": [
        {"key": "cycle_id",         "label": "Cycle",           "type": "select", "source": "cycles"},
        {"key": "department_id",    "label": "Department",      "type": "select", "source": "departments"},
        {"key": "status",           "label": "Status",          "type": "select", "options": [("pending","Pending"),("self_assessed","Self Assessed"),("reviewed","Reviewed"),("moderated","Moderated"),("approved","Approved")]},
    ],
    "transfer": [
        {"key": "transfer_type",    "label": "Transfer Type",   "type": "select", "options": [("transfer","Transfer"),("rotation","Rotation"),("secondment","Secondment"),("promotion","Promotion")]},
        {"key": "status",           "label": "Status",          "type": "select", "options": [("pending","Pending"),("approved","Approved"),("rejected","Rejected"),("completed","Completed")]},
        {"key": "department_id",    "label": "Department",      "type": "select", "source": "departments"},
    ],
    "manpower": [
        {"key": "department_id",    "label": "Department",      "type": "select", "source": "departments"},
        {"key": "financial_year",   "label": "Financial Year",  "type": "text"},
        {"key": "status",           "label": "Status",          "type": "select", "options": [("draft","Draft"),("submitted","Submitted"),("approved","Approved"),("rejected","Rejected")]},
    ],
    "trainee": [
        {"key": "program_id",       "label": "Program",         "type": "select", "source": "programs"},
        {"key": "department_id",    "label": "Department",      "type": "select", "source": "departments"},
        {"key": "status",           "label": "Status",          "type": "select", "options": [("active","Active"),("completed","Completed"),("terminated","Terminated")]},
        {"key": "trainee_type",     "label": "Type",            "type": "select", "options": [("graduate","Graduate Trainee"),("intern","Student Intern")]},
    ],
}


def generate_report(report_type, report_format, parameters=None):
    """
    Generate a report file.

    Args:
        report_type (str): One of REPORT_GENERATORS keys.
        report_format (str): 'csv', 'excel', or 'pdf'.
        parameters (dict): Filter parameters.

    Returns:
        tuple: (file_bytes, filename, content_type)
    """
    if parameters is None:
        parameters = {}

    extractor = REPORT_GENERATORS.get(report_type)
    if not extractor:
        raise ValueError(f"Unknown report type: {report_type}")

    headers, data = extractor(parameters)
    title = REPORT_LABELS.get(report_type, report_type.title())
    timestamp = date.today().isoformat()

    if report_format == "csv":
        content = generate_csv(headers, data)
        filename = f"{report_type}_{timestamp}.csv"
        content_type = "text/csv"
    elif report_format == "excel":
        content = generate_excel(headers, data, title)
        filename = f"{report_type}_{timestamp}.xlsx"
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif report_format == "pdf":
        content = generate_pdf(headers, data, title)
        filename = f"{report_type}_{timestamp}.pdf"
        content_type = "application/pdf"
    else:
        raise ValueError(f"Unsupported format: {report_format}")

    return content, filename, content_type